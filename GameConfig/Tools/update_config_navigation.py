# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:
    print("缺少 Python 依赖 openpyxl，请先安装：python -m pip install openpyxl", file=sys.stderr)
    raise SystemExit(1) from exc


GAME_CONFIG_DIR = Path(__file__).resolve().parents[1]
DEFAULT_DATA_DIR = GAME_CONFIG_DIR / "Datas"
DEFAULT_OUTPUT = GAME_CONFIG_DIR / "配置表导航.xlsx"
DEFAULT_REFERENCE = Path(r"D:\Unity\Unity_Project\data\xls\配置文件表.xlsm")
META_FILES = {"__tables__.xlsx", "__beans__.xlsx", "__enums__.xlsx"}


def value_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def is_marker(value) -> bool:
    return isinstance(value, str) and value.startswith("##")


def sheet_location(sheet_name: str) -> str:
    return f"'{sheet_name.replace(chr(39), chr(39) * 2)}'!A1"


def excel_text(value: str) -> str:
    return value.replace('"', '""')


def copy_reference_styles(reference: Path) -> dict[str, dict[str, object]]:
    if not reference.exists():
        return {}

    wb = load_workbook(reference, keep_vba=reference.suffix.lower() == ".xlsm", data_only=False)
    ws = wb["配置表目录"] if "配置表目录" in wb.sheetnames else wb.worksheets[0]

    def pack(cell_name: str) -> dict[str, object]:
        cell = ws[cell_name]
        return {
            "fill": copy(cell.fill),
            "font": copy(cell.font),
            "alignment": copy(cell.alignment),
            "border": copy(cell.border),
        }

    return {
        "header": pack("A1"),
        "folder": pack("A2"),
        "file": pack("B3"),
        "sheet": pack("C4"),
    }


def apply_style(cell, style: dict[str, object] | None):
    if not style:
        return
    for key, value in style.items():
        setattr(cell, key, copy(value))


def relative_excel_path(path: Path, base: Path) -> str:
    return str(path.relative_to(base)).replace("/", "\\")


def find_file(data_dir: Path, file_name: str) -> Path | None:
    for path in data_dir.rglob(file_name):
        if not path.name.startswith("~$"):
            return path
    return None


def parse_registry(data_dir: Path) -> dict[str, list[dict[str, str]]]:
    path = find_file(data_dir, "__tables__.xlsx")
    if path is None or not path.exists():
        return {}

    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    headers: list[str] = []
    for idx, row in enumerate(rows):
        if row and row[0] == "##var":
            header_idx = idx
            headers = [value_text(v) for v in row]
            break
    if header_idx is None:
        return {}

    col = {name: i for i, name in enumerate(headers) if name}

    def get(row, name: str, default: str = "") -> str:
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return default
        return value_text(row[idx]) or default

    registry: dict[str, list[dict[str, str]]] = {}
    for row in rows[header_idx + 1 :]:
        if not row or is_marker(row[0]) or not any(v not in (None, "") for v in row):
            continue
        input_value = get(row, "input")
        if not input_value:
            continue
        info = {
            "full_name": get(row, "full_name"),
            "value_type": get(row, "value_type"),
            "index": get(row, "index"),
            "mode": get(row, "mode", "map"),
            "group": get(row, "group"),
            "comment": get(row, "comment"),
            "tags": get(row, "tags"),
        }
        table_dir = path.parent
        for input_name in [part.strip() for part in input_value.split(",") if part.strip()]:
            registry.setdefault(input_name, []).append(info)
            candidate = table_dir / input_name
            if candidate.exists():
                registry.setdefault(relative_excel_path(candidate, data_dir), []).append(info)
    return registry


def extract_sheet_info(path: Path, data_dir: Path, registry: dict[str, list[dict[str, str]]]) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=False)
    data_relative = relative_excel_path(path, data_dir)
    registered_tables = registry.get(data_relative, []) or registry.get(path.name, [])
    result: list[dict[str, object]] = []

    for sheet_index, ws in enumerate(wb.worksheets, start=1):
        rows = list(ws.iter_rows(values_only=True))
        marker_rows: list[int] = []
        header_idx = None
        type_idx = None
        group_idx = None
        comment_idx = None

        for idx, row in enumerate(rows[:30]):
            first = row[0] if row else None
            if not is_marker(first):
                continue
            marker_rows.append(idx)
            if first == "##var" and header_idx is None:
                header_idx = idx
            elif first == "##type" and type_idx is None:
                type_idx = idx
            elif first == "##group" and group_idx is None:
                group_idx = idx
            elif first in ("##comment", "##") and comment_idx is None:
                comment_idx = idx

        fields = []
        export_field_count = 0
        if header_idx is not None:
            names = rows[header_idx]
            types = rows[type_idx] if type_idx is not None else []
            groups = rows[group_idx] if group_idx is not None else []
            comments = rows[comment_idx] if comment_idx is not None else []
            for col_idx in range(1, len(names)):
                name = value_text(names[col_idx])
                if not name:
                    continue
                ignored = name.startswith("##")
                if not ignored:
                    export_field_count += 1
                fields.append(
                    {
                        "name": name,
                        "type": value_text(types[col_idx]) if col_idx < len(types) else "",
                        "group": value_text(groups[col_idx]) if col_idx < len(groups) else "",
                        "comment": value_text(comments[col_idx]) if col_idx < len(comments) else "",
                        "ignored": ignored,
                        "column": col_idx + 1,
                    }
                )

        data_start = max(marker_rows) + 1 if marker_rows else 0
        data_rows = 0
        for row in rows[data_start:]:
            if not row or is_marker(row[0]):
                continue
            if any(v not in (None, "") for v in row):
                data_rows += 1

        registry_info: dict[str, str] = {}
        if len(registered_tables) == 1:
            registry_info = registered_tables[0]
        elif sheet_index - 1 < len(registered_tables):
            registry_info = registered_tables[sheet_index - 1]

        result.append(
            {
                "file": data_relative,
                "sheet": ws.title,
                "category": "Luban元表" if path.name in META_FILES else "业务配置表",
                "max_col": ws.max_column,
                "data_rows": data_rows,
                "field_count": export_field_count,
                "fields": fields,
                "registry": registry_info,
            }
        )
    return result


def add_link(cell, relative_target: str, sheet_name: str | None = None):
    display = value_text(cell.value)
    if sheet_name:
        link = f"{relative_target}#{sheet_location(sheet_name)}"
    else:
        link = relative_target
    cell.value = f'=HYPERLINK("{excel_text(link)}","{excel_text(display)}")'
    cell.style = "Hyperlink"


def style_header(ws, row: int, max_col: int, styles: dict[str, dict[str, object]]):
    for cell in ws[row]:
        if cell.column > max_col:
            continue
        apply_style(cell, styles.get("header"))
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def final_format(ws):
    ws.sheet_view.showGridLines = False
    widths = {
        "A": 16,
        "B": 24,
        "C": 22,
        "D": 20,
        "E": 18,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 34,
        "J": 14,
        "K": 12,
        "L": 28,
        "M": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border


def build_navigation(data_dir: Path, output: Path, reference: Path | None, include_meta: bool) -> tuple[int, int, int]:
    if not data_dir.exists():
        raise FileNotFoundError(f"配置表目录不存在：{data_dir}")

    styles = copy_reference_styles(reference) if reference else {}
    registry = parse_registry(data_dir)
    files = sorted(
        [
            path
            for path in data_dir.rglob("*.xlsx")
            if not path.name.startswith("~$")
            and path.resolve() != output.resolve()
            and (include_meta or path.name not in META_FILES)
        ],
        key=lambda p: (p.name not in META_FILES, relative_excel_path(p, data_dir)),
    )

    all_sheets = []
    for path in files:
        all_sheets.extend(extract_sheet_info(path, data_dir, registry))

    wb = Workbook()
    nav = wb.active
    nav.title = "配置表目录"
    detail = wb.create_sheet("表详情")
    field_index = wb.create_sheet("字段索引")
    doc = wb.create_sheet("doc")

    nav_headers = ["文件夹", "文件名", "工作表名", "Luban表名", "记录类型", "模式", "数据行数", "字段数", "说明"]
    nav.append(nav_headers)
    style_header(nav, 1, len(nav_headers), styles)
    nav.freeze_panes = "A2"
    nav.auto_filter.ref = "A1:I1"

    row = 2
    nav.cell(row, 1, f"{data_dir.name}\\")
    apply_style(nav.cell(row, 1), styles.get("folder"))
    nav.cell(row, 9, f"{data_dir.as_posix()} 下所有 Excel 配置表")
    row += 1

    for path in files:
        data_relative = relative_excel_path(path, data_dir)
        file_sheets = [item for item in all_sheets if item["file"] == data_relative]
        relative = str(path.relative_to(output.parent)).replace("/", "\\")
        nav.cell(row, 2, data_relative)
        add_link(nav.cell(row, 2), relative)
        nav.cell(row, 9, "Luban 元定义表" if path.name in META_FILES else "业务配置表")
        row += 1
        for info in file_sheets:
            registry_info = info["registry"]
            nav.cell(row, 3, info["sheet"])
            add_link(nav.cell(row, 3), relative, value_text(info["sheet"]))
            nav.cell(row, 4, registry_info.get("full_name", ""))
            nav.cell(row, 5, registry_info.get("value_type", ""))
            nav.cell(row, 6, registry_info.get("mode", ""))
            nav.cell(row, 7, info["data_rows"])
            nav.cell(row, 8, info["field_count"])
            nav.cell(row, 9, registry_info.get("comment", ""))
            row += 1

    detail_headers = [
        "分类",
        "文件名",
        "工作表名",
        "Luban表名",
        "记录类型",
        "索引",
        "模式",
        "分组",
        "数据行数",
        "导出字段数",
        "总列数",
        "注册说明",
        "Tags",
    ]
    detail.append(detail_headers)
    style_header(detail, 1, len(detail_headers), styles)
    for info in all_sheets:
        registry_info = info["registry"]
        detail.append(
            [
                info["category"],
                info["file"],
                info["sheet"],
                registry_info.get("full_name", ""),
                registry_info.get("value_type", ""),
                registry_info.get("index", ""),
                registry_info.get("mode", ""),
                registry_info.get("group", ""),
                info["data_rows"],
                info["field_count"],
                info["max_col"],
                registry_info.get("comment", ""),
                registry_info.get("tags", ""),
            ]
        )
        current = detail.max_row
        relative = str((data_dir / value_text(info["file"])).relative_to(output.parent)).replace("/", "\\")
        add_link(detail.cell(current, 2), relative)
        add_link(detail.cell(current, 3), relative, value_text(info["sheet"]))

    field_headers = ["文件名", "工作表名", "列号", "字段名", "类型", "分组", "说明", "注释列"]
    field_index.append(field_headers)
    style_header(field_index, 1, len(field_headers), styles)
    for info in all_sheets:
        relative = str((data_dir / value_text(info["file"])).relative_to(output.parent)).replace("/", "\\")
        for field in info["fields"]:
            field_index.append(
                [
                    info["file"],
                    info["sheet"],
                    field["column"],
                    field["name"],
                    field["type"],
                    field["group"],
                    field["comment"],
                    "是" if field["ignored"] else "",
                ]
            )
            current = field_index.max_row
            add_link(field_index.cell(current, 1), relative)
            add_link(field_index.cell(current, 2), relative, value_text(info["sheet"]))

    doc_rows = [
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["扫描目录", str(data_dir)],
        ["输出文件", str(output)],
        ["参考文件", str(reference) if reference else ""],
        ["导航结构", "参考 LeanNew 配置表目录：文件夹 -> 文件名 -> 工作表名，工作表名单元格链接到对应 Excel Sheet。"],
        ["文件格式", "本脚本生成 xlsx，无 VBA 宏；如果需要点击文件夹折叠，请基于 xlsm 宏模板另行扩展。"],
        ["数据行数规则", "从最后一个 Luban 标记行（##var/##type/##group/## 等）之后开始统计非空数据行。"],
        ["字段数规则", "字段名前缀为 ## 的注释列不计入导出字段数，但仍收录在 字段索引。"],
    ]
    for item in doc_rows:
        doc.append(item)
    style_header(doc, 1, 2, styles)

    for ws in (nav, detail, field_index, doc):
        final_format(ws)
    detail.freeze_panes = "A2"
    field_index.freeze_panes = "A2"

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    field_count = sum(len(info["fields"]) for info in all_sheets)
    return len(files), len(all_sheets), field_count


def parse_args():
    parser = argparse.ArgumentParser(description="更新 GameConfig 配置表目录导航。")
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR, help="配置表目录，默认 GameConfig/Datas")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="输出导航文件，默认 GameConfig/配置表导航.xlsx")
    parser.add_argument("--reference", type=Path, default=DEFAULT_REFERENCE, help="样式参考文件；不存在时使用内置样式")
    parser.add_argument("--no-reference", action="store_true", help="不读取参考文件样式")
    parser.add_argument("--exclude-meta", action="store_true", help="不把 __tables__/__beans__/__enums__ 加入导航")
    return parser.parse_args()


def main():
    args = parse_args()
    reference = None if args.no_reference else args.reference
    if reference and not reference.exists():
        print(f"参考文件不存在，改用内置样式：{reference}")
        reference = None

    file_count, sheet_count, field_count = build_navigation(
        data_dir=args.data_dir.resolve(),
        output=args.output.resolve(),
        reference=reference.resolve() if reference else None,
        include_meta=not args.exclude_meta,
    )
    print(f"已更新：{args.output.resolve()}")
    print(f"文件数：{file_count}，Sheet数：{sheet_count}，字段条目：{field_count}")


if __name__ == "__main__":
    main()
